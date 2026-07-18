import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from app.database import get_db
from app import models, auth

router = APIRouter(prefix="/reports", tags=["Reports"])


def _invoices_for_business(db: Session, business_id: str):
    return (
        db.query(models.Invoice)
        .filter(models.Invoice.business_id == business_id, models.Invoice.is_voided == False)  # noqa: E712
        .all()
    )


@router.get("/outstanding.xlsx")
def outstanding_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    invoices = [i for i in _invoices_for_business(db, current_user.business_id) if i.outstanding_amount > 0]

    wb = Workbook()
    ws = wb.active
    ws.title = "Outstanding"
    ws.append(["Invoice Number", "Retailer", "Invoice Date", "Amount", "Outstanding", "Status", "Due Date"])
    for inv in invoices:
        ws.append(
            [
                inv.invoice_number,
                inv.retailer.name if inv.retailer else "",
                str(inv.invoice_date),
                inv.amount,
                inv.outstanding_amount,
                inv.status.value,
                str(inv.due_date),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=outstanding_report.xlsx"},
    )


@router.get("/overdue.xlsx")
def overdue_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    invoices = [
        i
        for i in _invoices_for_business(db, current_user.business_id)
        if i.status == models.InvoiceStatus.overdue
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Overdue"
    ws.append(["Invoice Number", "Retailer", "Due Date", "Outstanding"])
    for inv in invoices:
        ws.append([inv.invoice_number, inv.retailer.name if inv.retailer else "", str(inv.due_date), inv.outstanding_amount])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=overdue_report.xlsx"},
    )


@router.get("/collection.xlsx")
def collection_report(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    payments = (
        db.query(models.Payment)
        .join(models.Invoice)
        .filter(models.Invoice.business_id == current_user.business_id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Collections"
    ws.append(["Invoice Number", "Payment Date", "Amount", "Delay Days", "Reference"])
    for p in payments:
        ws.append(
            [p.invoice.invoice_number, str(p.payment_date), p.amount, p.delay_days, p.reference_number or ""]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=collection_report.xlsx"},
    )
